import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from huobi.client.generic import GenericClient
from huobi.client.market import MarketClient
from huobi.constant import *
from huobi.utils import *
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from huobi.exception.huobi_api_exception import HuobiApiException
import telegram
from telegram import Bot
from apscheduler.schedulers.blocking import BlockingScheduler
import pytz
import time
import logging
import os

logging.basicConfig(level=logging.INFO)

TELEGRAM_BOT_TOKEN = ''
TELEGRAM_CHAT_ID = ''

# Set the ATR periods and candlestick interval
SHORT_ATR_PERIOD = 30  # Short-term ATR period
LONG_ATR_PERIOD = 288  # Long-term ATR period
CANDLESTICK_INTERVAL = CandlestickInterval.MIN5  # Candlestick interval
CANDLESTICK_INTERVAL_2 = CandlestickInterval.DAY1

output_directory = r"HTX_TOKEN_DATA"

# Known invalid symbols (you can update this list based on your findings)
invalid_symbols = set([
    "ltc3lusdt", "valueusdt", "btc3lusdt", "bpusdt", "antexusdt", "bagsusdt", "eth3lusdt", "pgalausdt",
    "tornusdt", "swrvusdt", "wndusdt", "bixusdt", "celusdt", "oldnormieusdt", "gbpusdt", "dacusdt",
    "lenausdt", "eos3susdt", "primateusdt", "wndrusdt", "seanusdt", "taousdt", "cubeusdt", "aeusdt",
    "dzoousdt", "itcusdt", "bitusdt", "mtloldusdt", "tincusdt", "titanusdt", "atsusdt", "ruffusdt",
    "chsbusdt", "mcusdt", "itausdt", "wildusdt", "xrtusdt", "anvusdt", "hoopusdt", "lolusdt", "hypeusdt",
    "socusdt", "ringusdt", "dieusdt", "warusdt", "mineusdt", "bch3lusdt", "fil3lusdt", "poolzusdt",
    "gcoinusdt", "htrusdt", "storeusdt", "lhbusdt", "dot2lusdt", "mapusdt", "xrp3susdt", "eth1susdt",
    "massusdt", "bhdusdt", "gofusdt", "tnbusdt", "emusdt", "onitusdt", "btmusdt", "tokeusdt", "fotausdt",
    "aegisusdt", "uni2susdt", "porusdt", "matchusdt", "edenusdt", "mcgusdt", "dot2susdt", "noiausdt",
    "pvtusdt", "nsureusdt", "irisusdt", "usnusdt", "vvsusdt", "qomusdt", "brtusdt", "lbausdt", "uuuusdt",
    "alt0usdt", "ftiusdt", "sdnusdt", "crousdt", "rethusdt", "zec3susdt", "koiusdt", "egtusdt", "creusdt",
    "ssxusdt", "poktusdt", "venusdt", "bch3susdt", "wozxusdt", "bxenusdt", "botusdt", "kcashusdt", "nbsusdt",
    "gearusdt", "radarusdt", "xrp3lusdt", "onstonusdt", "ogousdt", "wbtcusdt", "dhtusdt", "yeeusdt",
    "paradoxusdt", "stnusdt", "arixusdt", "actusdt", "upiusdt", "skmusdt", "vidyusdt", "zksusdt", "ektusdt",
    "gfusdt", "mplusdt", "seeleusdt", "eos3lusdt", "hcusdt", "foxusdt", "ltc3susdt", "slcusdt", "jumbousdt",
    "oceanusdt", "bchausdt", "btsusdt", "hydrousdt", "ctxc2xusdt", "dtausdt", "plausdt", "hitusdt", "skuusdt",
    "fsnusdt", "starlyusdt", "rifiusdt", "jennerusdt", "o3usdt", "1solusdt", "fildausdt", "hbcusdt", "egsusdt",
    "ncashusdt", "hptusdt", "indiusdt", "vempusdt", "muusdt", "btc1susdt", "octusdt", "kmausdt", "yamusdt",
    "nctusdt", "lunrusdt", "platousdt", "atpusdt", "letusdt", "plyusdt", "tdxusdt", "eurusdt", "eth3susdt",
    "zbcusdt", "dfausdt", "bsv3susdt", "spumeusdt", "cnnsusdt", "stakeusdt", "rndrusdt", "arknusdt",
    "ryomausdt", "mirusdt", "kaiusdt", "spsusdt", "revousdt", "sosusdt", "astusdt", "coolusdt", "wtcusdt",
    "galftusdt", "lxtusdt", "neiroethusdt", "agixusdt", "canusdt", "boousdt", "uipusdt", "mageusdt",
    "gnxusdt", "nhbtcusdt", "nuusdt", "chadusdt", "ioiusdt", "capousdt", "uni2lusdt", "talkusdt", "nasusdt",
    "btc3susdt", "htusdt", "efiusdt", "zntusdt", "mmfusdt", "lendusdt", "cmtusdt", "unicusdt", "fil3susdt",
    "xmxusdt", "ocnusdt", "abtusdt", "skebusdt", "gxcusdt", "zec3lusdt", "lmrusdt", "shitusdt", "mcousdt",
    "egameusdt", "ntusdt", "multiusdt", "nodeusdt", "dockusdt", "pearlusdt", "mdsusdt", "gstusdt", "fiuusdt",
    "invusdt", "screamusdt", "basebotusdt", "mtausdt", "rcccusdt", "tribeusdt", "loomusdt", "apnusdt",
    "yamv2usdt", "galusdt", "paiusdt", "sudousdt", "phcrusdt", "gvrusdt", "link3lusdt"
])

bot = Bot(token=TELEGRAM_BOT_TOKEN)

def fetch_all_symbols():
    generic_client = GenericClient()
    symbols = generic_client.get_exchange_symbols()
    symbols_usdt = [symbol.symbol.lower() for symbol in symbols if symbol.quote_currency == 'usdt']
    return symbols_usdt

def is_valid_symbol(symbol):
    """Quickly validate if the symbol is valid by attempting to fetch a single candlestick."""
    market_client = MarketClient()
    try:
        list_obj = market_client.get_candlestick(symbol, CANDLESTICK_INTERVAL, 1)
        return len(list_obj) > 0
    except HuobiApiException:
        return False

def fetch_historical_data(symbol, size, period=CANDLESTICK_INTERVAL_2):
    market_client = MarketClient()
    try:
        list_obj = market_client.get_candlestick(symbol, period, size)
        if list_obj and len(list_obj) > 0:
            data = []
            for candlestick in list_obj:
                data.append({
                    'high': candlestick.high,
                    'low': candlestick.low,
                    'close': candlestick.close,
                    'vol': candlestick.vol,
                    'open': candlestick.open
                })
            return data
        else:
            print(f"No 'data' found in response for symbol {symbol}")
            return None
    except HuobiApiException as e:
        print(f"Skipping symbol {symbol} due to API exception: {e}")
        return None

def calculate_atr(ohlcv_data, atr_period):
    # Convert to DataFrame for easier calculations
    df = pd.DataFrame(ohlcv_data)
    df['high'] = df['high']
    df['low'] = df['low']
    df['close'] = df['close']
    
    # Calculate True Range (TR) as a percentage of the closing price
    df['previous_close'] = df['close'].shift(1)
    df['tr_percentage'] = df.apply(lambda row: max(row['high'] - row['low'],
                                                   abs(row['high'] - row['previous_close']),
                                                   abs(row['low'] - row['previous_close'])) / row['close'] * 100, axis=1)
    
    # Calculate ATR using a rolling window of TR percentages
    df['atr_percentage'] = df['tr_percentage'].rolling(window=atr_period).mean()
    
    # Return the latest ATR value as a percentage
    return df['atr_percentage'].iloc[-1]

import os  # Import the os module for handling file paths

# Define the output directory where you want to save the Excel files
output_directory = r"C:\Users\erikm\OneDrive\Desktop\DEV\HTX_TOKEN_DATA"

def record_data():
    # Fetch all symbols (trading pairs) that have USDT as the quote currency
    symbols = fetch_all_symbols()
    
    processed_data = []
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    date_str = datetime.now().strftime('%Y-%m-%d')
    time_str = datetime.now().strftime('%H-%M-%S')  # Add time to ensure uniqueness
    
    for symbol in symbols:
        if not is_valid_symbol(symbol):
            print(f"Skipping invalid symbol {symbol}")
            continue
        
        # Fetch data for the longer ATR period (to cover both short and long ATRs)
        historical_data_long = fetch_historical_data(symbol, size=LONG_ATR_PERIOD)
        
        if historical_data_long:
            # Calculate short-term and long-term ATR percentages
            short_atr_value = calculate_atr(historical_data_long[-SHORT_ATR_PERIOD:], SHORT_ATR_PERIOD)
            long_atr_value = calculate_atr(historical_data_long, LONG_ATR_PERIOD)
            
            # Calculate Consolidation Factor using both ATR values
            consolidation_factor = short_atr_value / long_atr_value if long_atr_value != 0 else None
            
            # Calculate Percent Difference
            latest_data = historical_data_long[0]
            percent_diff = ((latest_data['high'] - latest_data['low']) / latest_data['low']) * 100 if latest_data['low'] != 0 else 0
            
            processed_data.append({
                'timestamp': timestamp,
                'symbol': symbol,
                'open': latest_data['open'],
                'high': latest_data['high'],
                'low': latest_data['low'],
                'close': latest_data['close'],
                'volume': latest_data['vol'],
                'percent_diff': percent_diff,
                'short_atr_percentage': short_atr_value,
                'long_atr_percentage': long_atr_value,
                'consolidation_factor': consolidation_factor
            })
        else:
            print(f"Skipping symbol {symbol} due to data issues.")
    
    # Create a DataFrame to store the data
    df = pd.DataFrame(processed_data)
    
    # Generate the file name with current date and time
    file_name = f'htx_token_data_{date_str}_{time_str}.xlsx'
    
    # Combine the directory and file name to get the full file path
    file_path = os.path.join(output_directory, file_name)
    
    # Ensure the output directory exists; if not, create it
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    # Save data to an Excel file in the specified directory
    df.to_excel(file_path, index=False)
    
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Apply conditional formatting to the 'short_atr_percentage' column
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    
    atr_column = 'J'  # Assuming 'short_atr_percentage' is in column J
    ws.conditional_formatting.add(
        f'{atr_column}2:{atr_column}{len(df) + 1}', 
        CellIsRule(operator='lessThan', formula=['10'], stopIfTrue=True, fill=green_fill)
    )
    
    # Save the workbook with the conditional formatting
    wb.save(file_path)
    print(f"Recorded data at {timestamp}")
    return file_path

def send_telegram_message(file_path, retries=3, delay=5):
    """Send the Excel file to the Telegram bot with retry mechanism."""
    for attempt in range(retries):
        try:
            with open(file_path, 'rb') as file:
                bot.send_document(chat_id=TELEGRAM_CHAT_ID, document=file, caption="Here is the HTX Token Data Report.")
            print("Telegram message sent successfully")
            return  # Exit the function if the message is sent successfully
        except Exception as e:
            logging.error(f"Error sending message to Telegram: {e}")
            if attempt < retries - 1:
                print(f"Retrying in {delay} seconds...")
                time.sleep(delay)  # Wait before retrying
            else:
                print("Max retries reached. Failed to send message.")

def job():
    file_path = record_data()
    if file_path:
        #send_email(file_path)
        send_telegram_message(file_path)

scheduler = BlockingScheduler()
timezone = pytz.timezone('UTC')
scheduler.add_job(job, 'interval', hours=1, timezone=timezone)

# Run the job immediately for testing
if __name__ == "__main__":
    job()
    scheduler.start()
